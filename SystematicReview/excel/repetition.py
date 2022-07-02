import csv
import pandas as pd
import xlrd
from openpyxl import Workbook


def repetition():
    filepath = 'file/JCR.xlsx'
    data = xlrd.open_workbook(filepath)
    sh = data.sheet_by_name('Sheet1')
    lent = sh.nrows
    list_jrc = []
    list_jrc_isnn = []
    for i in range(1, lent):
        # print(i,"  ",sh.row_values(i))
        da = sh.row_values(i)
        if da[1] not in list_jrc_isnn:
            list_jrc_isnn.append(da[1])
            list_jrc.append(da)
    print("list_jrc ", len(list_jrc), len(list_jrc_isnn))

    wb = Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    ws = wb.active
    ws['A1'] = 'Journal name'
    ws['B1'] = 'ISSN'
    ws['C1'] = 'eISSN'
    ws['D1'] = 'Category'
    ws['E1'] = 'Total Citations'
    ws['F1'] = '2020 JIF'
    ws['G1'] = 'JIF Quartile'
    ws['H1'] = 'SCI'
    ws['I1'] = '2020JCI'
    ws['J1'] = '% of OA Gold'

    count2 = 0
    for row2 in list_jrc:
        # print(row2,i)
        ws.append(row2)
        count2 += 1
    wb.save("file/JCR_norep.xlsx")


def JCR_noRep():
    filepath = 'file/JCR_norep.xlsx'
    data = xlrd.open_workbook(filepath)
    sh = data.sheet_by_name('Sheet')
    lent = sh.nrows
    list_jrc = []
    list_jrc_isnn = []
    for i in range(1, lent):
        # print(i,"  ",sh.row_values(i))
        da = sh.row_values(i)
        if da[7] != '':
            list_jrc.append(da)
    print("list_jrc ", len(list_jrc), len(list_jrc_isnn))

    wb = Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    ws = wb.active
    ws['A1'] = 'Journal name'
    ws['B1'] = 'ISSN'
    ws['C1'] = 'eISSN'
    ws['D1'] = 'Category'
    ws['E1'] = 'Total Citations'
    ws['F1'] = '2020 JIF'
    ws['G1'] = 'JIF Quartile'
    ws['H1'] = 'SCI'
    ws['I1'] = '2020JCI'
    ws['J1'] = '% of OA Gold'

    count2 = 0
    for row2 in list_jrc:
        ws.append(row2)
        count2 += 1
    wb.save("file_out/JCR_norep2.xlsx")


# 加载CCF
def loadCCF():
    filepath = 'file_out/CCF.xlsx'
    data = xlrd.open_workbook(filepath)
    sh = data.sheet_by_name('International')
    lent = sh.nrows
    list_cff = []
    list_cff_title = []
    print("length :", lent - 1)

    for i in range(1, lent):
        # print(i,"  ",sh.row_values(i))
        da = sh.row_values(i)
        # title = da[2].replace(" ", "")  # 去除所有空格
        title = ''.join(filter(str.isalpha, da[2])).lower()  # 只保留英文过滤,且小写
        if title not in list_cff_title:
            list_cff_title.append(title)
            list_cff.append(da)
        else:
            print(da[2])  # 打印文件中重复的期刊名
    print("list_cff ", len(list_cff), len(list_cff_title))
    return list_cff, list_cff_title


def loadSCI():
    filepath = 'file_out/JCR_norep2.xlsx'
    data = xlrd.open_workbook(filepath)
    sh = data.sheet_by_name('Sheet')
    lent = sh.nrows
    list_cff = []
    list_cff_title = []
    print("length :", lent - 1)

    for i in range(1, lent):
        # print(i,"  ",sh.row_values(i))
        da = sh.row_values(i)
        # title = da[0].replace(" ", "")  # 去除所有空格
        title = ''.join(filter(str.isalpha, da[0])).lower()  # 只保留英文过滤,且小写
        if title not in list_cff_title:
            list_cff_title.append(title)
            list_cff.append(da)
        else:
            print(da[0])  # 打印文件中重复的期刊名
    print("list_JCR ", len(list_cff), len(list_cff_title))
    return list_cff, list_cff_title


def fusion():
    ccf, ccf_t = loadCCF()
    sci, sci_t = loadSCI()
    #
    ccf_f = []
    for c in ccf:
        title_c = ''.join(filter(str.isalpha, c[2])).lower()  # 只保留英文过滤,且小写
        for s in sci:
            title_s = ''.join(filter(str.isalpha, s[0])).lower()  # 只保留英文过滤,且小写
            if title_c == title_s:
                c.append(s[6])
                c.append(s[7])
                continue

    wb = Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    ws = wb.active
    ws['A1'] = '序号'
    ws['B1'] = '期刊（会议）简称'
    ws['C1'] = '刊物（会议）全称'
    ws['D1'] = '出版社'
    ws['E1'] = '网址'
    ws['F1'] = '期刊or会议'
    ws['G1'] = '类别'
    ws['H1'] = '学科领域'
    ws['I1'] = 'JIF Quartile'
    ws['J1'] = 'SCI中科院分区'

    count2 = 0
    for row in ccf:
        ws.append(row)
        count2 += 1
    wb.save("file_out/CCF_SCI.xlsx")


def classfy():
    filepath = 'file_out/CCF_SCI2.xlsx'
    data = xlrd.open_workbook(filepath)
    sh = data.sheet_by_name('Sheet')
    lent = sh.nrows
    ccf_sci1 = []  # sci分区不为空
    ccf_sci2 = []  # sci分区为空
    for i in range(1, lent):
        # print(i,"  ",sh.row_values(i))
        da = sh.row_values(i)
        if da[8] != '':
            da[9] = str(int(da[9])) + '区'
            ccf_sci1.append(da)
        else:
            ccf_sci2.append(da)

    wb = Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    ws = wb.active
    ws['A1'] = '序号'
    ws['B1'] = '期刊（会议）简称'
    ws['C1'] = '刊物（会议）全称'
    ws['D1'] = '出版社'
    ws['E1'] = '网址'
    ws['F1'] = '期刊or会议'
    ws['G1'] = '类别'
    ws['H1'] = '学科领域'
    ws['I1'] = 'JIF Quartile'
    ws['J1'] = 'SCI中科院分区'

    for row in ccf_sci1:
        ws.append(row)
    for row in ccf_sci2:
        ws.append(row)
    wb.save("file_out/CCF_SCI.xlsx")


def addqu():
    return


list_jcr = []
filepath = 'file_out/JCR_norep2.xlsx'
data = xlrd.open_workbook(filepath)
sh = data.sheet_by_name('Sheet')
lent = sh.nrows
for i in range(1, lent):
    da = sh.row_values(i)
    da[5] = str(int(da[5])) + '区'
    list_jcr.append(da)

wb = Workbook()
# 获取当前活跃的sheet，默认是第一个sheet
ws = wb.active
ws['A1'] = 'Journal name'
ws['B1'] = 'ISSN'
ws['C1'] = 'eISSN'
ws['D1'] = 'Total Citations'
ws['E1'] = '2020 JIF'
ws['F1'] = 'SCI'
ws['G1'] = 'JIF Quartile'
ws['H1'] = '2020JCI'
ws['I1'] = '% of OA Gold'

for row2 in list_jcr:
    ws.append(row2)
wb.save("file_out/JCR_final.xlsx")

# a = 4.0
# print(a)
# a = int(a)
# print(a)

# str1 = 'IEEE Transactions on Computer-Aided Design of Integrated Circuits And System '
# str2 = 'IEEE TRANSACTIONS ON COMPUTER-AIDED DESIGN OF INTEGRATED CIRCUITS AND SYSTEMS'
