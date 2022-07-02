import csv
import pandas as pd
import xlrd
from openpyxl import Workbook

def scopus():
    filepath1 = 'file2/scopus.csv'
    list_scopus = []
    list_scopus_title = []
    with open(filepath1,'r',encoding='UTF-8') as f:
        reader = csv.reader(f)
        # header = next(reader)
        # print(header)
        for row in reader:
            #print(row[2])
            title = ''.join(filter(str.isalpha,row[2])).lower()  #只保留英文过滤
            list_scopus_title.append(title)
            list_scopus.append([row[12], row[3], row[0], row[2], row[16]])
    print(len(list_scopus_title))
    list_scopus_title = list_scopus_title[1:]
    list_scopus = list_scopus[1:]
    print("scopus论文数量 ",len(list_scopus_title))

    # filepath2 = 'file2/scopus_2.csv'
    # list_scopus2 = []
    # list_scopus_title2 = []
    # with open(filepath2,'r',encoding='UTF-8') as f:
    #     reader = csv.reader(f)
    #     # header = next(reader)
    #     # print(header)
    #     for row in reader:
    #         #print(row[2])
    #         title = ''.join(filter(str.isalpha,row[2])).lower()  #只保留英文过滤
    #         list_scopus_title2.append(title)
    #         list_scopus2.append([row[12], row[3], row[0], row[2], row[16]])
    # print(len(list_scopus_title2))
    # list_scopus_title2 = list_scopus_title2[1:]
    # list_scopus2 = list_scopus2[1:]
    # print("scopus论文数量 ",len(list_scopus_title2))
    #
    # list_scopus3 = []
    # list_scopus_title3 = []
    # for r in list_scopus2:
    #     title = ''.join(filter(str.isalpha,r[3])).lower()  #只保留英文过滤
    #     if title not in list_scopus_title:
    #         list_scopus.append(r)
    #         list_scopus_title.append(title)
    #
    print("总论文数量 ",len(list_scopus))
    return list_scopus,list_scopus_title

def webofscience():
    filepath3 = 'file2/savedrecs.xls'
    #filepath4 = 'file2/savedrecs1001_.xls'
    # df = pd.read_excel(filepath2,sheet_name='savedrecs')
    # data = df.head()
    # print(data)
    # for p in data:
    #     print(p)
    list_web = []
    list_web_title = []
    data = xlrd.open_workbook(filepath3)
    sh = data.sheet_by_name('savedrecs')
    lent = sh.nrows

    # data2 = xlrd.open_workbook(filepath4)
    # sh2 = data2.sheet_by_name('savedrecs')
    # lent2 = sh2.nrows

    # print(sh.nrows)#有效数据行数

    # da = sh.row_values(0)
    # print(da)
    # list_web.append([da[28],da[33],da[1],da[9],da[34]])

    for i in range(1, lent):
        # print(i,"  ",sh.row_values(i))
        da = sh.row_values(i)
        title = ''.join(filter(str.isalpha, da[9])).lower()  # 只保留英文过滤
        list_web_title.append(title)
        list_web.append([da[28], da[33], da[1], da[9], da[34]])
    # for i in range(1, lent2):
    #     # print(i,"  ",sh.row_values(i))
    #     da = sh2.row_values(i)
    #     title = ''.join(filter(str.isalpha, da[9])).lower()  # 只保留英文过滤
    #     list_web_title.append(title)
    #     list_web.append([da[28], da[33], da[1], da[9], da[34]])

    print("web of science ", len(list_web),type(list_web))
    return list_web,list_web_title


def ieee():
    filepath1 = 'file2/ieee.csv'
    list_ieee = []
    list_ieee_title = []
    with open(filepath1,'r',encoding='UTF-8') as f:
        reader = csv.reader(f)
        # header = next(reader)
        # print(header)
        # title:0 Author:1 Year:5  abbtract:10  DOI:13
        #13 5 1 0 10
        for row in reader:
            #print(row[2])
            title = ''.join(filter(str.isalpha, row[0])).lower()  # 只保留英文过滤,且小写
            list_ieee_title.append(title)
            list_ieee.append([row[13], row[5], row[1], row[0], row[10]])
    print(len(list_ieee_title))
    list_ieee_title = list_ieee_title[1:]
    list_ieee = list_ieee[1:]
    print("ieee论文数量 ",len(list_ieee_title))
    return list_ieee,list_ieee_title

def duplicated_remove(list_in):
    list_r = []
    list_r_t = []
    for row in list_in:
        title = ''.join(filter(str.isalpha, row[3])).lower()  # 只保留英文过滤
        if title not in list_r_t:
            list_r_t.append(title)
            list_r.append(row)
    return list_r,list_r_t

def totalpaper(scopus,scopus_t,webos,webos_t,ieee):

    for w in webos:
        title = ''.join(filter(str.isalpha, w[3])).lower()  # 只保留英文过滤
        if title not in scopus_t:
            scopus.append(w)
            scopus_t.append(title)
    for ie in ieee:
        title = ''.join(filter(str.isalpha, ie[3])).lower()  # 只保留英文过滤
        if title not in scopus_t:
            scopus.append(ie)
            scopus_t.append(title)
    return scopus,scopus_t

list_s1,list_s_t1 = scopus()
print("scopus论文总数1",len(list_s1),len(list_s_t1))
list_s,list_s_t = duplicated_remove(list_s1)
print("scopus论文总数2",len(list_s),len(list_s_t))

list_w1,list_w_t1 = webofscience()
print("web of science 论文总数1",len(list_w1),len(list_w_t1))
list_w,list_w_t = duplicated_remove(list_w1)
print("web of science 论文总数2",len(list_w),len(list_w_t))

list_ie1,list_ie_t1 = ieee()
print("ieee论文数量 ", len(list_ie1),len(list_ie_t1))
list_ie,list_ie_t = duplicated_remove(list_ie1)
print("去重后ieee论文数量 ", len(list_ie),len(list_ie_t))

list_p,list_p_t = totalpaper(list_s,list_s_t,list_w,list_w_t,list_ie)
print("总paper数  去重前",len(list_ie1)+len(list_w1)+len(list_s1),"  去重后 ",len(list_p))
print("重复数量 ",len(list_ie1)+len(list_w1)+len(list_s1)-len(list_p))

wb = Workbook()
# 获取当前活跃的sheet，默认是第一个sheet
ws = wb.active
ws['A1'] = 'DOI'
ws['B1'] = 'Publication Year'
ws['C1'] = 'Authors'
ws['D1'] = 'Article Title'
ws['E1'] = 'Abstract'

count2 = 0
for row2 in list_p:
    # print(row2,i)
    ws.append(row2)
    count2 += 1
wb.save("file_out/gaze_xlsx.xlsx")

# count2 = 0
# with open(filepath5, 'a+', newline='',encoding='UTF-8') as f:
#     writer = csv.writer(f)
#     writer.writerow(['DOI', 'Publication Year','Authors', 'Article Title', 'Abstract'])
#     for row in list_s:
#         writer.writerow(row)
#         count2 +=1
#
#     for row2 in list_w:
#         #print(row2,i)
#         title = ''.join(filter(str.isalpha,row2[3])).lower()  #只保留英文过滤
#         if title not in list_s_t:
#             writer.writerow(row2)
#             count2+=1


#
#
# print("list_web_title",len(list_web_title),type(list_web_title))
# print("web of science 论文数量",len(list_web))
#

#
# print('Scopus&WebOfScience total_num=',count2)
#
# count = 0
# for title in list_web_title:
#     if title not in list_scopus_title:
#         count += 1
# print('重复论文 count = ',len(list_web_title)-count)
#

