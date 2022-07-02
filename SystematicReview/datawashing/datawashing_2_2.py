import csv
import pandas as pd
import xlrd
from openpyxl import Workbook

def duplicated_remove(list_in):
    list_r = []
    list_r_t = []
    for row in list_in:
        title = ''.join(filter(str.isalpha, row[3])).lower()  # 只保留英文过滤
        if title not in list_r_t:
            list_r_t.append(title)
            list_r.append(row)
    return list_r,list_r_t

def ieee():
    filepath1 = 'file2/ieee.csv'
    list_ieee = []
    list_ieee_title = []
    with open(filepath1,'r',encoding='UTF-8') as f:
        reader = csv.reader(f)
        # header = next(reader)
        # print(header)
        # title:0 Author:1 Year:5  abbtract:10  DOI:13
        #13 5 1 0 10
        for row in reader:
            #print(row[2])
            title = ''.join(filter(str.isalpha, row[0])).lower()  # 只保留英文过滤,且小写
            list_ieee_title.append(title)
            list_ieee.append([row[13], row[5], row[1], row[0], row[10]])
    print(len(list_ieee_title))
    list_ieee_title = list_ieee_title[1:]
    list_ieee = list_ieee[1:]
    print("ieee论文数量 ",len(list_ieee_title))
    return list_ieee,list_ieee_title

list_ie1,list_ie_t1 = ieee()
print("ieee论文数量 ", len(list_ie1),len(list_ie_t1))

list_ie,list_ie_t = duplicated_remove(list_ie1)
print("去重后ieee论文数量 ", len(list_ie),len(list_ie_t))

filepath3 = 'file2/gaze.xlsx'

list_ws = []
list_ws_title = []
data = xlrd.open_workbook(filepath3)
sh = data.sheet_by_name('Sheet')
lent = sh.nrows


for i in range(1, lent):
    # print(i,"  ",sh.row_values(i))
    da = sh.row_values(i)
    title = ''.join(filter(str.isalpha, da[3])).lower()  # 只保留英文过滤,且小写
    list_ws.append(da)
    list_ws_title.append(title)

    # if(da[5]>0):
    #     title = ''.join(filter(str.isalpha, da[3])).lower()  # 只保留英文过滤,且小写
    #     list_ws.append(da)
    #     list_ws_title.append(title)

print("scopus and webofscience ",len(list_ws),len(list_ws_title))

for ie in list_ie:
    title = ''.join(filter(str.isalpha, ie[3])).lower()  # 只保留英文过滤,且小写
    if title not in list_ws_title:
        list_ws_title.append(title)
        list_ws.append(ie)
print("三个库 总数",len(list_ws))

wb = Workbook()
# 获取当前活跃的sheet，默认是第一个sheet
ws = wb.active

ws['A1'] = 'DOI'
ws['B1'] = 'Publication Year'
ws['C1'] = 'Authors'
ws['D1'] = 'Article Title'
ws['E1'] = 'Abstract'
ws['F1'] = 'Screening(Include=1,Exclude=0,maybe=0)'
ws['G1']='Full text checked(yes=1,no=0)'
ws['H1'] = 'driver=2,HCI=3,VR=4,other(pupil)=1'
for row in list_ws:
    ws.append(row)
wb.save('file_out/gaze_xlsx.xlsx')