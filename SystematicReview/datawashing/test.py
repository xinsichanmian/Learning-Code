from openpyxl import Workbook
# 创建一个Workbook对象

import csv
import pandas as pd
import xlrd

filepath2 = 'file_out/gaze_xlsx.xlsx'
# df = pd.read_excel(filepath2,sheet_name='savedrecs')
# data = df.head()
# print(data)
# for p in data:
#     print(p)
list_web = []
list_web_title = []
data = xlrd.open_workbook(filepath2)
sh = data.sheet_by_name('Sheet')
lent = sh.nrows
# print(sh.nrows)#有效数据行数

# da = sh.row_values(0)
# print(da)
# list_web.append([da[28],da[33],da[1],da[9],da[34]])

for i in range(1,lent):
    #print(i,"  ",sh.row_values(i))
    da = sh.row_values(i)
    title = ''.join(filter(str.isalpha,da[3])).lower()  #只保留英文过滤
    list_web_title.append(title)
    list_web.append(da)

print("list_web_title",len(list_web_title),type(list_web_title))
print("web of science 论文数量",len(list_web))

list_web2 = []
list_web_title2 = []
for row in list_web:
    title = ''.join(filter(str.isalpha, row[3])).lower()  # 只保留英文过滤
    if title not in list_web_title2:
        list_web_title2.append(title)
        list_web2.append(row)

print("list_web_title2",len(list_web_title2),type(list_web_title2))
print("论文数量2",len(list_web2))



wb = Workbook()
# 获取当前活跃的sheet，默认是第一个sheet
ws = wb.active
ws['A1'] = 'DOI'
ws['B1'] = 'Publication Year'
ws['C1'] = 'Authors'
ws['D1'] = 'Article Title'
ws['E1'] = 'Abstract'

count2 = 0
for row in list_web2:
    ws.append(row)
    count2 += 1

wb.save("file_out/gaze_xlsx2.xlsx")
print("论文总数 ",count2)
