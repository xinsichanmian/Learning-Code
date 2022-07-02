import csv
import pandas as pd
import xlrd
from openpyxl import Workbook


filepath3 = 'file2/gaze_xlsx.xlsx'
# filepath4 = 'file2/savedrecs1001_.xls'
# df = pd.read_excel(filepath2,sheet_name='savedrecs')
# data = df.head()
# print(data)
# for p in data:
#     print(p)
list_web = []
list_web_title = []
data = xlrd.open_workbook(filepath3)
sh = data.sheet_by_name('Sheet')
lent = sh.nrows

# data2 = xlrd.open_workbook(filepath4)
# sh2 = data2.sheet_by_name('savedrecs')
# lent2 = sh2.nrows

# print(sh.nrows)#有效数据行数

# da = sh.row_values(0)
# print(da)
# list_web.append([da[28],da[33],da[1],da[9],da[34]])

for i in range(1, lent):
    # print(i,"  ",sh.row_values(i))
    da = sh.row_values(i)
    if da[5]>0:
        list_web.append(da)
print("经过摘要和标题筛选，剩余 ",len(list_web))
# for i in range(1, lent2):
#     # print(i,"  ",sh.row_values(i))
#     da = sh2.row_values(i)
#     title = ''.join(filter(str.isalpha, da[9])).lower()  # 只保留英文过滤
#     list_web_title.append(title)
#     list_web.append([da[28], da[33], da[1], da[9], da[34]])
wb = Workbook()
# 获取当前活跃的sheet，默认是第一个sheet
ws = wb.active

ws['A1'] = 'DOI'
ws['B1'] = 'Publication Year'
ws['C1'] = 'Authors'
ws['D1'] = 'Article Title'
ws['E1'] = 'Abstract'
ws['F1'] = 'Screening(Include=1,Exclude=0,maybe=0)'
ws['G1']='Full text checked(yes=1,no=0)'
ws['H1'] = 'driver=2,HCI=3,VR=4,other(pupil)=1'
for row in list_web:
    ws.append(row)
wb.save('file_out/gaze_xlsx2.xlsx')
